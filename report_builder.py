import calendar
from datetime import date, datetime
from pathlib import Path

import holidays as pyholidays
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from models import DailyManualEntry, DailyReport, VehicleStore


def _kr_holidays(year: int):
    return pyholidays.country_holidays("KR", years=[year])


def _working_days(year: int, month: int) -> list[date]:
    _, last_day = calendar.monthrange(year, month)
    kr_holidays = _kr_holidays(year)
    days = []
    for day in range(1, last_day + 1):
        d = date(year, month, day)
        if d.weekday() >= 5:
            continue
        if d in kr_holidays:
            continue
        days.append(d)
    return days


def _month_range(year: int, month: int) -> tuple[date, date]:
    _, last_day = calendar.monthrange(year, month)
    return date(year, month, 1), date(year, month, last_day)


def _merged_daily_value(report: DailyReport | None, manual: DailyManualEntry | None) -> dict:
    values = {
        "passenger_name": None,
        "start_time": report.start_time if report else None,
        "end_time": report.end_time if report else None,
        "odometer_start": report.odometer_start if report else None,
        "odometer_end": report.odometer_end if report else None,
        "distance_km": report.distance_km if report else None,
    }

    if manual:
        values["passenger_name"] = manual.passenger_name
        if manual.start_time is not None:
            values["start_time"] = manual.start_time
        if manual.end_time is not None:
            values["end_time"] = manual.end_time
        if manual.odometer_start is not None:
            values["odometer_start"] = manual.odometer_start
        if manual.odometer_end is not None:
            values["odometer_end"] = manual.odometer_end
        if manual.distance_km is not None:
            values["distance_km"] = manual.distance_km

    return values


def _build_monthly_rows(db, car_id: str, year: int, month: int) -> tuple[str, list[list]]:
    vehicle = db.query(VehicleStore).filter(VehicleStore.car_id == car_id).first()
    car_type_text = ""
    if vehicle:
        car_type_text = vehicle.car_sellname or vehicle.car_name or ""

    start_of_month, end_of_month = _month_range(year, month)
    days = _working_days(year, month)

    reports = (
        db.query(DailyReport)
        .filter(
            DailyReport.car_id == car_id,
            DailyReport.drive_date >= start_of_month,
            DailyReport.drive_date <= end_of_month,
        )
        .all()
    )
    manuals = (
        db.query(DailyManualEntry)
        .filter(
            DailyManualEntry.car_id == car_id,
            DailyManualEntry.drive_date >= start_of_month,
            DailyManualEntry.drive_date <= end_of_month,
        )
        .all()
    )

    report_by_date = {r.drive_date: r for r in reports}
    manual_by_date = {m.drive_date: m for m in manuals}

    rows = []
    for drive_date in days:
        report = report_by_date.get(drive_date)
        manual = manual_by_date.get(drive_date)
        merged = _merged_daily_value(report, manual)

        distance_km = merged.get("distance_km")
        has_manual_values = bool(
            (manual and manual.passenger_name)
            or (manual and manual.start_time)
            or (manual and manual.end_time)
            or (manual and manual.odometer_start is not None)
            or (manual and manual.odometer_end is not None)
            or (manual and manual.distance_km is not None)
        )

        is_zero_day = distance_km in (None, 0)
        if is_zero_day and not has_manual_values:
            values = [
                drive_date.strftime("%Y-%m-%d"),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        else:
            values = [
                drive_date.strftime("%Y-%m-%d"),
                merged.get("passenger_name") or "",
                merged.get("start_time") or "",
                merged.get("odometer_start") if merged.get("odometer_start") is not None else "",
                merged.get("end_time") or "",
                merged.get("odometer_end") if merged.get("odometer_end") is not None else "",
                merged.get("distance_km") if merged.get("distance_km") is not None else "",
                "",
            ]
        rows.append(values)

    return car_type_text, rows


def generate_monthly_report_xlsx(db, car_id: str, year: int, month: int, output_dir: str = "outputs") -> str:
    car_type_text, rows = _build_monthly_rows(db, car_id, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "운행일지"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    header_font = Font(name="맑은 고딕", size=11, bold=True)
    body_font = Font(name="맑은 고딕", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"차량 운행일지 - {year}년 ({month})월"
    ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True)
    ws["A1"].alignment = center

    ws["A2"] = "반"
    ws["B2"] = "30반"
    ws["C2"] = "차종"
    ws["D2"] = car_type_text
    ws["E2"] = "차량ID"
    ws["F2"] = car_id

    for cell in ("A2", "B2", "C2", "D2", "E2", "F2"):
        ws[cell].font = body_font
        ws[cell].alignment = center if cell in {"A2", "C2", "E2"} else left

    headers = ["일자", "탑승자", "출발시간", "출발km", "도착시간", "도착km", "운행거리", "사고유무"]
    row_start = 4
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row_start, column=idx, value=title)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    row = row_start + 1
    for values in rows:
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = body_font
            cell.alignment = center
            cell.border = border
        row += 1

    footer_row = row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    ws.cell(row=footer_row, column=1, value="담당/팀장/과장 결재선 및 운전자 서명란은 공란")
    ws.cell(row=footer_row, column=1).font = Font(name="맑은 고딕", size=9)

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    filename = f"vehicle_log_{car_id}_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = output / filename
    wb.save(file_path)

    return str(file_path.resolve())


def _resolve_pdf_font_name() -> str:
    malgun_path = Path("C:/Windows/Fonts/malgun.ttf")
    if malgun_path.exists():
        font_name = "MalgunGothic"
        if font_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(font_name, str(malgun_path)))
        return font_name
    return "Helvetica"


def generate_monthly_report_pdf(db, car_id: str, year: int, month: int, output_dir: str = "outputs") -> str:
    car_type_text, rows = _build_monthly_rows(db, car_id, year, month)
    font_name = _resolve_pdf_font_name()

    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    filename = f"vehicle_log_{car_id}_{year}_{month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    file_path = output / filename

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleKR",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=14,
        leading=18,
        alignment=1,
    )
    body_style = ParagraphStyle(
        name="BodyKR",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=12,
    )

    elements = []
    elements.append(Paragraph(f"차량 운행일지 - {year}년 ({month})월", title_style))
    elements.append(Spacer(1, 8))

    meta_data = [["반", "30반", "차종", car_type_text, "차량ID", car_id]]
    meta_table = Table(meta_data, colWidths=[34, 70, 34, 120, 45, 180])
    meta_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(meta_table)
    elements.append(Spacer(1, 8))

    headers = ["일자", "탑승자", "출발시간", "출발km", "도착시간", "도착km", "운행거리", "사고유무"]
    table_data = [headers] + rows
    main_table = Table(
        table_data,
        colWidths=[68, 58, 58, 58, 58, 58, 58, 58],
        repeatRows=1,
    )
    main_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(main_table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("담당/팀장/과장 결재선 및 운전자 서명란은 공란", body_style))

    doc.build(elements)
    return str(file_path.resolve())
